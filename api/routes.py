"""FastAPI routes."""

from __future__ import annotations
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
import os, math

from engine.package import PackageResult, recommend_package
from engine.slab import verify_slab_rule, slab, volumetric_kg, billable_kg
from engine.weight import WeightResult, estimate_weight, validate_on_labeled_set
from engine.db import ch_query

router = APIRouter()

EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "orders.xlsx")
_SEJALIMPEX = "764c0f86-cb84-432f-a77f-b8f6cfaaed4f"


# ---------------------------------------------------------------------------
# Shared models
# ---------------------------------------------------------------------------

class WeightRequest(BaseModel):
    sku: str
    sku_name: str
    quantity: int = 1
    client_id: str
    declared_weight_kg: Optional[float] = None
    package_length_cm: Optional[float] = None
    package_width_cm: Optional[float] = None
    package_height_cm: Optional[float] = None


class WeightResponse(BaseModel):
    predicted_dead_weight_kg: float
    confidence: float
    drives_slab: bool
    basis: str
    target_slab: float
    notes: list[str]


class PackageRequest(BaseModel):
    sku: str
    client_id: str
    applied_package_id: str
    billing_mode: str = "box"
    use_vision: bool = False


class PackageResponse(BaseModel):
    decision: str
    current_package_id: str
    current_slab: float
    target_slab: float
    recommended_package_id: Optional[str] = None
    suggested_new_dims_cm: Optional[dict] = None
    confidence: float
    evidence: dict
    reason: str


# ---------------------------------------------------------------------------
# Analyze-order — the main endpoint the frontend calls
# ---------------------------------------------------------------------------

class OrderRequest(BaseModel):
    # Order identity
    sku: str
    sku_name: str
    quantity: int = 1
    client_id: str

    # What the seller declared
    applied_weight_kg: Optional[float] = None   # seller's declared billable
    applied_package_id: Optional[str] = None

    # Package dims (if known)
    package_length_cm: Optional[float] = None
    package_width_cm: Optional[float] = None
    package_height_cm: Optional[float] = None


class WeightIssue(BaseModel):
    severity: str          # "ok" | "warning" | "critical"
    title: str
    detail: str
    suggestion: str


class OrderAnalysis(BaseModel):
    # Weight engine
    predicted_dead_weight_kg: float
    predicted_slab: float
    applied_slab: Optional[float]
    weight_confidence: float
    weight_basis: str
    weight_issues: List[WeightIssue]

    # Package engine (if package_id given)
    package_decision: Optional[str]        # "keep" | "switch" | "create" | None
    package_reason: Optional[str]
    recommended_package_id: Optional[str]
    suggested_new_dims_cm: Optional[dict]
    package_confidence: Optional[float]

    # Summary
    overall_status: str                 # "ok" | "warning" | "critical"
    summary: str


def _build_weight_issues(
    predicted_kg: float,
    predicted_slab: float,
    applied_weight_kg: Optional[float],
    pkg_l: Optional[float],
    pkg_w: Optional[float],
    pkg_h: Optional[float],
    basis: str,
    notes: List[str],
    sku: str = "",
    client_id: str = "",
    applied_package_id: Optional[str] = None,
) -> List[WeightIssue]:
    issues: List[WeightIssue] = []

    # ── Pull historical context from ClickHouse ──────────────────────────────
    hist_context = _get_historical_context(sku, client_id, applied_package_id)

    # ── No declared weight ───────────────────────────────────────────────────
    if applied_weight_kg is None or applied_weight_kg <= 0:
        issues.append(WeightIssue(
            severity="warning",
            title="No applied weight declared",
            detail="Seller has not declared a billable weight for this order.",
            suggestion=f"Declare ≥ {predicted_kg:.3f} kg (predicted dead weight) → expected slab {predicted_slab} kg.",
        ))
        if hist_context:
            issues.append(hist_context)
        return issues

    applied_slab = slab(applied_weight_kg)

    # Volumetric from dims
    if pkg_l and pkg_w and pkg_h:
        vol = volumetric_kg(pkg_l, pkg_w, pkg_h)
        bill = billable_kg(applied_weight_kg, vol)
        applied_slab = slab(bill)

    slab_diff = predicted_slab - applied_slab

    if abs(slab_diff) < 0.01:
        issues.append(WeightIssue(
            severity="ok",
            title="Weight slab matches prediction",
            detail=f"Applied weight {applied_weight_kg:.3f} kg → slab {applied_slab} kg matches predicted slab {predicted_slab} kg.",
            suggestion="No change needed.",
        ))
        if hist_context:
            issues.append(hist_context)
        return issues

    # ── Slab mismatch ────────────────────────────────────────────────────────
    direction = "under-declared" if slab_diff > 0 else "over-declared"
    slabs_off = abs(slab_diff) / 0.5
    severity = "critical" if abs(slab_diff) >= 1.0 else "warning"

    # Explain WHY the weight is wrong
    if basis == "sku_history":
        detail = (
            f"Based on {notes[0] if notes else 'sorter history'}, the median re-weighed "
            f"weight is {predicted_kg:.3f} kg (slab {predicted_slab} kg). "
            f"You declared {applied_weight_kg:.3f} kg → slab {applied_slab} kg. "
            f"This is {direction} by {slabs_off:.0f} slab{'s' if slabs_off > 1 else ''}."
        )
        suggestion = (
            f"Update declared weight to ≥ {predicted_kg:.3f} kg. "
            f"Historical sorter data consistently shows this SKU weighs more than declared. "
            f"Failing to correct this will result in carrier weight discrepancy charges of "
            f"typically ₹30–80 per shipment."
        )
    else:
        detail = (
            f"Product name analysis estimates {predicted_kg:.3f} kg "
            f"(slab {predicted_slab} kg) based on item categories and pack quantities. "
            f"You declared {applied_weight_kg:.3f} kg → slab {applied_slab} kg. "
            f"This is {direction} by {slabs_off:.0f} slab{'s' if slabs_off > 1 else ''}."
        )
        suggestion = (
            f"Declare weight as ≥ {predicted_kg:.3f} kg to land in slab {predicted_slab} kg. "
            f"Under-declaring means the carrier will re-weigh and charge for the higher slab, "
            f"creating a weight discrepancy."
        )

    issues.append(WeightIssue(
        severity=severity,
        title=f"Weight {direction} — {slabs_off:.0f} slab{'s' if slabs_off > 1 else ''} off",
        detail=detail,
        suggestion=suggestion,
    ))

    if basis == "name_parse":
        issues.append(WeightIssue(
            severity="warning",
            title="Estimate based on product name (no sorter history for this SKU)",
            detail="No past sorter re-weighs found for this SKU. Weight estimated from product category and pack quantities.",
            suggestion="Once this shipment is scanned at the sorter, confidence will improve for future orders.",
        ))

    if hist_context:
        issues.append(hist_context)

    return issues


def _get_historical_context(sku: str, client_id: str, package_id: Optional[str]) -> Optional[WeightIssue]:
    """Pull sorter history stats for this SKU/package and return a context issue."""
    if not client_id:
        return None
    try:
        # Package-level history: how often does this package land in which slab?
        if package_id:
            df = ch_query("""
                SELECT
                    count() AS n,
                    countIf(max_dead_vol_sorter > 0) AS with_sorter,
                    median(max_dead_vol_sorter) AS med_billable,
                    countIf(sorter_slab > weight_slab_shipfast) AS higher_slab_count
                FROM shipfast_weight_discrepancy.shipfast_weight_discrepancy
                WHERE client_id = {cid:String}
                  AND package_id = {pid:String}
            """, {"cid": client_id, "pid": package_id})

            if not df.empty and int(df["with_sorter"].iloc[0] or 0) >= 5:
                n_sorter = int(df["with_sorter"].iloc[0])
                med_bill = float(df["med_billable"].iloc[0] or 0)
                higher = int(df["higher_slab_count"].iloc[0] or 0)
                pct = round(100 * higher / n_sorter) if n_sorter else 0
                typical_slab = slab(med_bill) if med_bill > 0 else None

                if pct >= 50 and typical_slab:
                    return WeightIssue(
                        severity="critical",
                        title=f"⚠ Historical risk: {pct}% of past orders with this package got a higher carrier slab",
                        detail=(
                            f"Out of {n_sorter} sorter-scanned shipments using this package, "
                            f"{higher} ({pct}%) were billed at a higher slab by the carrier. "
                            f"The median sorter billable weight is {med_bill:.3f} kg → slab {typical_slab} kg."
                        ),
                        suggestion=(
                            f"If the current order is not corrected, there is a high probability "
                            f"of a weight discrepancy charge. Recommended action: declare weight "
                            f"≥ {med_bill:.2f} kg or switch to a package that lands in slab {typical_slab} kg."
                        ),
                    )
    except Exception:
        pass
    return None


@router.post("/analyze-order", response_model=OrderAnalysis)
def analyze_order(req: OrderRequest):
    # --- Weight engine ---
    w: WeightResult = estimate_weight(
        sku=req.sku,
        sku_name=req.sku_name,
        quantity=req.quantity,
        client_id=req.client_id,
        declared_weight_kg=req.applied_weight_kg,
        pkg_l=req.package_length_cm,
        pkg_w=req.package_width_cm,
        pkg_h=req.package_height_cm,
    )

    # Applied slab from declared weight
    applied_slab: Optional[float] = None
    if req.applied_weight_kg and req.applied_weight_kg > 0:
        if req.package_length_cm and req.package_width_cm and req.package_height_cm:
            vol = volumetric_kg(req.package_length_cm, req.package_width_cm, req.package_height_cm)
            applied_slab = slab(billable_kg(req.applied_weight_kg, vol))
        else:
            applied_slab = slab(req.applied_weight_kg)

    weight_issues = _build_weight_issues(
        predicted_kg=w.predicted_dead_weight_kg,
        predicted_slab=w.target_slab,
        applied_weight_kg=req.applied_weight_kg,
        pkg_l=req.package_length_cm,
        pkg_w=req.package_width_cm,
        pkg_h=req.package_height_cm,
        basis=w.basis,
        notes=w.notes,
        sku=req.sku,
        client_id=req.client_id,
        applied_package_id=req.applied_package_id,
    )

    # --- Package engine (optional) ---
    pkg_decision = pkg_reason = rec_pkg_id = None
    suggested_dims = None
    pkg_confidence = None

    if req.applied_package_id:
        p: PackageResult = recommend_package(
            sku=req.sku,
            client_id=req.client_id,
            applied_package_id=req.applied_package_id,
            use_vision=False,
        )
        pkg_decision = p.decision
        pkg_reason = p.reason
        rec_pkg_id = p.recommended_package_id
        suggested_dims = p.suggested_new_dims_cm
        pkg_confidence = p.confidence

    # --- Overall status ---
    severities = [i.severity for i in weight_issues]
    if "critical" in severities or pkg_decision in ("switch", "create"):
        overall_status = "critical"
    elif "warning" in severities:
        overall_status = "warning"
    else:
        overall_status = "ok"

    # --- Summary sentence ---
    if overall_status == "ok":
        summary = f"Order looks correct. Predicted slab {w.target_slab} kg matches applied weight."
    else:
        parts = []
        for issue in weight_issues:
            if issue.severity in ("warning", "critical"):
                parts.append(issue.title)
        if pkg_decision == "switch":
            parts.append(f"Switch to package {rec_pkg_id}")
        elif pkg_decision == "create":
            parts.append("Create a new package to match the correct slab")
        summary = ". ".join(parts) + "." if parts else "Review required."

    return OrderAnalysis(
        predicted_dead_weight_kg=w.predicted_dead_weight_kg,
        predicted_slab=w.target_slab,
        applied_slab=applied_slab,
        weight_confidence=w.confidence,
        weight_basis=w.basis,
        weight_issues=weight_issues,
        package_decision=pkg_decision,
        package_reason=pkg_reason,
        recommended_package_id=rec_pkg_id,
        suggested_new_dims_cm=suggested_dims,
        package_confidence=pkg_confidence,
        overall_status=overall_status,
        summary=summary,
    )


# ---------------------------------------------------------------------------
# Existing routes
# ---------------------------------------------------------------------------

@router.post("/weight", response_model=WeightResponse)
def predict_weight(req: WeightRequest):
    result: WeightResult = estimate_weight(
        sku=req.sku,
        sku_name=req.sku_name,
        quantity=req.quantity,
        client_id=req.client_id,
        declared_weight_kg=req.declared_weight_kg,
        pkg_l=req.package_length_cm,
        pkg_w=req.package_width_cm,
        pkg_h=req.package_height_cm,
    )
    return WeightResponse(
        predicted_dead_weight_kg=result.predicted_dead_weight_kg,
        confidence=result.confidence,
        drives_slab=result.drives_slab,
        basis=result.basis,
        target_slab=result.target_slab,
        notes=result.notes,
    )


@router.post("/package", response_model=PackageResponse)
def recommend(req: PackageRequest):
    result: PackageResult = recommend_package(
        sku=req.sku,
        client_id=req.client_id,
        applied_package_id=req.applied_package_id,
        billing_mode=req.billing_mode,
        use_vision=req.use_vision,
    )
    return PackageResponse(
        decision=result.decision,
        current_package_id=result.current_package_id,
        current_slab=result.current_slab,
        target_slab=result.target_slab,
        recommended_package_id=result.recommended_package_id,
        suggested_new_dims_cm=result.suggested_new_dims_cm,
        confidence=result.confidence,
        evidence=result.evidence,
        reason=result.reason,
    )


# ---------------------------------------------------------------------------
# Package catalogue (for dropdown)
# ---------------------------------------------------------------------------

@router.get("/packages/{client_id}")
def list_packages(client_id: str):
    """Return all package IDs used by this client, with sorter-derived dims."""
    df = ch_query("""
        SELECT
            package_id                       AS id,
            count()                          AS shipments,
            any(package_type_seller)         AS pkg_type,
            median(min_sorter_length)        AS med_l,
            median(min_sorter_width)         AS med_w,
            median(min_sorter_height)        AS med_h,
            median(max_dead_vol_sorter)      AS med_billable,
            countIf(max_dead_vol_sorter > 0) AS with_sorter
        FROM shipfast_weight_discrepancy.shipfast_weight_discrepancy
        WHERE client_id = {cid:String}
          AND notEmpty(toString(package_id))
        GROUP BY package_id
        ORDER BY shipments DESC
    """, {"cid": client_id})

    if df.empty:
        return {"packages": []}

    packages = []
    for _, row in df.iterrows():
        med_l = float(row["med_l"]) if row["med_l"] and not _isnan(row["med_l"]) else None
        med_w = float(row["med_w"]) if row["med_w"] and not _isnan(row["med_w"]) else None
        med_h = float(row["med_h"]) if row["med_h"] and not _isnan(row["med_h"]) else None
        med_bill = float(row["med_billable"]) if row["med_billable"] and not _isnan(row["med_billable"]) else None
        pkg_slab = round(slab(med_bill), 1) if med_bill else None
        label = _package_label(str(row["id"]), int(row["shipments"]), med_l, med_w, med_h, pkg_slab)
        packages.append({
            "id": str(row["id"]),
            "shipments": int(row["shipments"]),
            "pkg_type": str(row["pkg_type"] or "box"),
            "med_l": med_l,
            "med_w": med_w,
            "med_h": med_h,
            "med_billable": med_bill,
            "typical_slab": pkg_slab,
            "label": label,
        })
    return {"packages": packages}


def _isnan(v) -> bool:
    try:
        return math.isnan(float(v))
    except Exception:
        return False


def _package_label(pkg_id: str, shipments: int, l, w, h, typical_slab) -> str:
    """Human-readable package label without showing the raw UUID."""
    if l and w and h:
        vol = float(l) * float(w) * float(h)
        if vol <= 3000:     size = "XS Box"
        elif vol <= 6000:   size = "S Box"
        elif vol <= 15000:  size = "M Box"
        elif vol <= 30000:  size = "L Box"
        elif vol <= 60000:  size = "XL Box"
        else:               size = "XXL Box"
        dims = f"{int(float(l))}×{int(float(w))}×{int(float(h))} cm"
        slab_str = f"~{typical_slab} kg slab" if typical_slab else ""
        parts = [p for p in [dims, slab_str, f"{shipments} shipments"] if p]
        return f"{size} ({', '.join(parts)})"
    else:
        return f"Box · {shipments} shipments · ID {pkg_id[:8]}…"


# ---------------------------------------------------------------------------
# Seeded orders from Excel (pre-existing history to display)
# ---------------------------------------------------------------------------

@router.get("/seeded-orders")
def seeded_orders(client_id: str = _SEJALIMPEX, limit: int = 50):
    """Return orders from the uploaded Excel file with pre-computed analysis."""
    try:
        import openpyxl
    except ImportError:
        return {"orders": []}

    xl = _find_excel()
    if not xl:
        return {"orders": []}

    try:
        wb = openpyxl.load_workbook(xl, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        key_cols = [
            "AWB", "order_id", "sku", "sku_name", "quantity", "package_id",
            "dead_weight_shipfast", "applied_weight_shipfast", "weight_slab_shipfast",
            "min_sorter_weight", "min_sorter_length", "min_sorter_width", "min_sorter_height",
            "max_dead_vol_sorter", "sorter_slab", "discrepancy_type_ai", "slab_change",
            "package_type_seller", "min_sorter_image_link", "carrier_name",
        ]
        idx = {c: headers.index(c) for c in key_cols if c in headers}

        orders = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if i >= limit * 2:
                break
            r = {c: row[idx[c]] for c in idx if c in idx}

            awb = str(r.get("AWB", "") or "")
            sku = str(r.get("sku", "") or "")
            sku_name = str(r.get("sku_name", "") or "")
            qty = int(r.get("quantity") or 1)
            pkg_id = str(r.get("package_id", "") or "")
            applied_wt = _f(r.get("applied_weight_shipfast"))
            declared_slab = _f(r.get("weight_slab_shipfast"))
            sorter_wt = _f(r.get("min_sorter_weight"))
            sorter_l = _f(r.get("min_sorter_length"))
            sorter_w = _f(r.get("min_sorter_width"))
            sorter_h = _f(r.get("min_sorter_height"))
            max_billable = _f(r.get("max_dead_vol_sorter"))
            sorter_slab_val = _f(r.get("sorter_slab"))
            discrepancy_type = str(r.get("discrepancy_type_ai", "") or "")
            slab_change = str(r.get("slab_change", "") or "")
            image_link = str(r.get("min_sorter_image_link", "") or "")
            carrier = str(r.get("carrier_name", "") or "")

            if not sku_name:
                continue

            # Determine analysis from Excel columns
            has_sorter = sorter_wt is not None and sorter_wt > 0
            if has_sorter and max_billable:
                actual_slab = sorter_slab_val or slab(max_billable)
            else:
                actual_slab = None

            # Build issues
            issues = _build_excel_issues(
                applied_wt=applied_wt,
                declared_slab=declared_slab,
                actual_slab=actual_slab,
                sorter_wt=sorter_wt,
                max_billable=max_billable,
                discrepancy_type=discrepancy_type,
                slab_change=slab_change,
                pkg_id=pkg_id,
                carrier=carrier,
                sku_name=sku_name,
            )

            worst = "ok"
            for iss in issues:
                if iss["severity"] == "critical":
                    worst = "critical"
                    break
                if iss["severity"] == "warning":
                    worst = "warning"

            orders.append({
                "id": awb or str(i),
                "awb": awb,
                "order_id": str(r.get("order_id", "") or ""),
                "sku": sku,
                "sku_name": sku_name,
                "quantity": qty,
                "package_id": pkg_id,
                "applied_weight_kg": applied_wt,
                "declared_slab": declared_slab,
                "sorter_weight": sorter_wt,
                "sorter_dims": [sorter_l, sorter_w, sorter_h] if sorter_l else None,
                "actual_slab": actual_slab,
                "max_billable": max_billable,
                "carrier": carrier,
                "sorter_image": image_link if image_link else None,
                "issues": issues,
                "overall_status": worst,
                "source": "excel",
            })

            if len(orders) >= limit:
                break

        return {"orders": orders, "total": len(orders)}
    except Exception as e:
        import logging
        logging.warning("seeded_orders error: %s", e)
        return {"orders": [], "error": str(e)}


def _f(v) -> Optional[float]:
    try:
        x = float(v)
        return x if not math.isnan(x) else None
    except Exception:
        return None


def _find_excel() -> Optional[str]:
    candidates = [
        EXCEL_PATH,
        os.path.expanduser("~/Downloads/weight_discrepancy___v3_2026-06-19T17_55_07.535173525+05_30.xlsx"),
        os.path.expanduser("~/Downloads/weight_discrepancy.xlsx"),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


def _build_excel_issues(
    applied_wt, declared_slab, actual_slab, sorter_wt, max_billable,
    discrepancy_type, slab_change, pkg_id, carrier, sku_name
) -> List[dict]:
    issues = []

    if actual_slab is not None and declared_slab is not None:
        slab_diff = actual_slab - declared_slab
        if abs(slab_diff) < 0.01:
            issues.append({
                "severity": "ok",
                "title": "Weight slab matches sorter",
                "detail": f"Declared slab {declared_slab} kg matches sorter-measured slab {actual_slab} kg.",
                "suggestion": "No weight change needed.",
            })
        elif slab_diff > 0:
            n_slabs = round(slab_diff / 0.5)
            issues.append({
                "severity": "critical" if slab_diff >= 1.0 else "warning",
                "title": f"Under-declared by {n_slabs} slab{'s' if n_slabs > 1 else ''}",
                "detail": (
                    f"You declared {declared_slab} kg slab, but the sorter measured "
                    f"billable weight {max_billable:.3f} kg → slab {actual_slab} kg. "
                    f"Dead weight at sorter: {sorter_wt:.3f} kg. "
                    f"Carrier: {carrier}."
                ),
                "suggestion": (
                    f"Increase declared weight to at least {max_billable:.2f} kg to reach "
                    f"slab {actual_slab} kg. Current under-declaration will result in a "
                    f"weight discrepancy charge from the carrier."
                ),
            })
            if "Higher slab applied by carrier" in slab_change:
                issues.append({
                    "severity": "critical",
                    "title": "Carrier has already applied higher slab",
                    "detail": f"Carrier ({carrier}) re-weighed and billed at {actual_slab} kg slab. {discrepancy_type}.",
                    "suggestion": "Either raise a dispute with evidence or accept and correct declared weight for future orders.",
                })
    elif declared_slab is not None and actual_slab is None:
        issues.append({
            "severity": "warning",
            "title": "No sorter data available",
            "detail": f"Declared slab is {declared_slab} kg but no sorter re-weigh available to verify.",
            "suggestion": "Weight will be verified when carrier scans the parcel.",
        })

    # Package check
    if pkg_id == "835c88e1-dde9-4758-acd4-74cdfae25953":
        issues.append({
            "severity": "warning",
            "title": "Default package applied to all orders",
            "detail": (
                "Package 835c88e1 (20×15×8 cm, dead_weight=0) is the default package "
                "applied to all orders. Historically 95%+ of shipments with this package "
                "are re-measured at a higher slab by the sorter."
            ),
            "suggestion": (
                "Review whether this package accurately reflects the shipment size. "
                "For orders with actual sorter dims > 20×15×8 cm, consider creating "
                "a correctly-sized package."
            ),
        })

    return issues


# ---------------------------------------------------------------------------
# Discrepancy insights (grouped recommendations for WD tab)
# ---------------------------------------------------------------------------

@router.get("/discrepancy-insights")
def discrepancy_insights(client_id: str = _SEJALIMPEX):
    """
    Analyse the Excel + ClickHouse data and return grouped root-cause insights
    with actionable recommendations to reduce weight discrepancies.
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed"}

    xl = _find_excel()
    if not xl:
        return {"error": "Excel file not found"}

    try:
        wb = openpyxl.load_workbook(xl, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        def hi(n): return headers.index(n) if n in headers else -1

        i_pkg     = hi("package_id")
        i_app_wt  = hi("applied_weight_shipfast")
        i_dec_sl  = hi("weight_slab_shipfast")
        i_sort_sl = hi("sorter_slab")
        i_sort_wt = hi("min_sorter_weight")
        i_max_bill= hi("max_dead_vol_sorter")
        i_disc    = hi("discrepancy_type_ai")
        i_sc      = hi("slab_change")
        i_carrier = hi("carrier_name")
        i_dw      = hi("dead_weight_shipfast")
        i_wd_ded  = hi("weight_discrepancy_charges_deducted")
        i_sku     = hi("sku_name")
        i_sort_l  = hi("min_sorter_length")
        i_sort_w  = hi("min_sorter_width")
        i_sort_h  = hi("min_sorter_height")

        total = 0; with_sorter = 0; higher_slab = 0
        zero_dead_wt = 0; default_pkg_orders = 0
        charges_total = 0.0; major_disc = 0; minor_disc = 0
        carrier_counts: dict = {}
        carrier_charges: dict = {}
        # Track orders where sorter dims > package dims (package too small)
        pkg_too_small = 0
        # Slab difference distribution
        slab_diffs: list = []

        DEFAULT_PKG = "835c88e1-dde9-4758-acd4-74cdfae25953"

        for row in ws.iter_rows(min_row=2, values_only=True):
            total += 1
            pkg        = str(row[i_pkg] or "") if i_pkg >= 0 else ""
            sorter_sl  = _f(row[i_sort_sl]) if i_sort_sl >= 0 else None
            declared_sl= _f(row[i_dec_sl])  if i_dec_sl >= 0 else None
            sorter_wt  = _f(row[i_sort_wt]) if i_sort_wt >= 0 else None
            max_bill   = _f(row[i_max_bill]) if i_max_bill >= 0 else None
            dw         = _f(row[i_dw])       if i_dw >= 0 else None
            carrier    = str(row[i_carrier] or "") if i_carrier >= 0 else ""
            disc       = str(row[i_disc] or "")    if i_disc >= 0 else ""
            sc         = str(row[i_sc] or "")      if i_sc >= 0 else ""
            wd_ded     = _f(row[i_wd_ded])          if i_wd_ded >= 0 else None
            sl         = _f(row[i_sort_l])           if i_sort_l >= 0 else None
            sw         = _f(row[i_sort_w])           if i_sort_w >= 0 else None
            sh         = _f(row[i_sort_h])           if i_sort_h >= 0 else None

            if sorter_wt and sorter_wt > 0:
                with_sorter += 1
            if sorter_sl and declared_sl and sorter_sl > declared_sl + 0.01:
                higher_slab += 1
                if sorter_sl and declared_sl:
                    slab_diffs.append(round(sorter_sl - declared_sl, 1))
            if not dw or dw == 0:
                zero_dead_wt += 1
            if pkg == DEFAULT_PKG:
                default_pkg_orders += 1
            if "Major" in disc:
                major_disc += 1
            elif "Minor" in disc:
                minor_disc += 1
            if wd_ded:
                charges_total += wd_ded
            if carrier:
                carrier_counts[carrier] = carrier_counts.get(carrier, 0) + 1
                carrier_charges[carrier] = carrier_charges.get(carrier, 0.0) + (wd_ded or 0)
            # Package too small: sorter dims > 22×17×10
            if sl and sw and sh and sl > 22 and pkg == DEFAULT_PKG:
                pkg_too_small += 1

        top_carrier = max(carrier_counts, key=carrier_counts.get) if carrier_counts else "Delhivery"
        top_carrier_pct = round(100 * carrier_counts.get(top_carrier, 0) / total) if total else 0
        top_carrier_charges = carrier_charges.get(top_carrier, 0.0)

        avg_slab_diff = round(sum(slab_diffs) / len(slab_diffs), 2) if slab_diffs else 0
        higher_pct = round(100 * higher_slab / with_sorter) if with_sorter else 0

        # Build grouped insight cards
        groups = [
            {
                "id": "zero_dead_weight",
                "severity": "critical",
                "title": "Dead weight declared as 0 on virtually all orders",
                "affected": zero_dead_wt,
                "affected_pct": round(100 * zero_dead_wt / total) if total else 0,
                "impact": f"When dead weight is 0, billing defaults to volumetric only. The sorter re-weighs the actual parcel and charges the correct slab, creating a discrepancy every time the dead weight exceeds the volumetric.",
                "recommendation": "Declare actual product dead weight for each order. For hair accessories, typical weights are 50–500 g depending on pack quantity. Use the weight engine to auto-predict.",
                "fix_effort": "Medium",
                "potential_savings": f"Prevents ~{higher_slab} future discrepancy charges",
            },
            {
                "id": "default_package",
                "severity": "critical",
                "title": f"99%+ orders use one default package (20×15×8 cm) regardless of actual product size",
                "affected": default_pkg_orders,
                "affected_pct": round(100 * default_pkg_orders / total) if total else 0,
                "impact": (
                    f"Package 835c88e1 (20×15×8 cm, dead weight 0) has volumetric = 0.48 kg → declared slab 0.5 kg. "
                    f"But the sorter consistently measures parcels at {avg_slab_diff + 0.5:.1f}+ kg slab. "
                    f"{pkg_too_small} of these orders had sorter dims larger than the declared package, "
                    f"proving the package registration doesn't match the actual box used."
                ),
                "recommendation": (
                    "Create size-specific packages (S/M/L/XL) that reflect actual box dimensions. "
                    "At minimum, update the default package dead weight to reflect actual tare (~100 g). "
                    "Ideally: for multi-item orders (≥5 items), switch to a larger registered package."
                ),
                "fix_effort": "Low",
                "potential_savings": f"Could eliminate {round(higher_pct)}% of slab mismatches",
            },
            {
                "id": "carrier_slab_higher",
                "severity": "warning",
                "title": f"{higher_slab} shipments had carrier apply a higher slab than declared",
                "affected": higher_slab,
                "affected_pct": higher_pct,
                "impact": (
                    f"{top_carrier} is the primary carrier ({top_carrier_pct}% of volume) and has charged "
                    f"₹{top_carrier_charges:,.0f} in weight discrepancy fees. "
                    f"Average slab difference: {avg_slab_diff} kg per affected shipment."
                ),
                "recommendation": (
                    "For all future orders: run the weight engine before shipping to predict the correct slab. "
                    "For past discrepancies marked 'New': review and raise disputes where the sorter image "
                    "shows dimensions matching your declared package (within tolerance)."
                ),
                "fix_effort": "High",
                "potential_savings": f"₹{charges_total:,.0f} already charged; prevent recurrence going forward",
            },
            {
                "id": "major_minor_disc",
                "severity": "warning",
                "title": f"{major_disc} major + {minor_disc} minor discrepancies recorded",
                "affected": major_disc + minor_disc,
                "affected_pct": round(100 * (major_disc + minor_disc) / total) if total else 0,
                "impact": (
                    f"Major discrepancies (≥1 slab difference) are harder to dispute and more expensive. "
                    f"Minor discrepancies (0.5 slab difference) are often borderline cases where better "
                    f"weight declaration would have prevented the charge entirely."
                ),
                "recommendation": (
                    "For major discrepancies: verify with sorter image; if parcel dimensions match declared "
                    "package, raise a dispute with photographic evidence. "
                    "For minor discrepancies: update declared weight upward by 0.1–0.2 kg to avoid the slab boundary."
                ),
                "fix_effort": "Medium",
                "potential_savings": f"Dispute resolution could recover a portion of ₹{charges_total:,.0f} charged",
            },
        ]

        return {
            "summary": {
                "total_orders": total,
                "with_sorter_data": with_sorter,
                "higher_slab_count": higher_slab,
                "higher_slab_pct": higher_pct,
                "total_wd_charges": round(charges_total, 2),
                "major_discrepancies": major_disc,
                "minor_discrepancies": minor_disc,
                "zero_dead_weight_orders": zero_dead_wt,
                "default_package_orders": default_pkg_orders,
                "top_carrier": top_carrier,
                "top_carrier_orders": carrier_counts.get(top_carrier, 0),
                "top_carrier_charges": round(top_carrier_charges, 2),
            },
            "groups": groups,
        }

    except Exception as e:
        import logging, traceback
        logging.warning("discrepancy_insights error: %s\n%s", e, traceback.format_exc())
        return {"error": str(e)}


# ---------------------------------------------------------------------------
# Verify slab rule
# ---------------------------------------------------------------------------

@router.get("/verify-slab-rule")
def verify_slab(client_id: str, limit: int = 5000):
    df = ch_query("""
        SELECT dead_weight_shipfast, volumetric_weight_shipfast,
               applied_weight_shipfast, weight_slab_shipfast
        FROM shipfast_weight_discrepancy.shipfast_weight_discrepancy
        WHERE client_id = {cid:String} AND weight_slab_shipfast > 0
        LIMIT {lim:Int32}
    """, {"cid": client_id, "lim": limit})
    return verify_slab_rule(df)


@router.get("/backtest/weight")
def backtest_weight(client_id: str, limit: int = 500):
    df = validate_on_labeled_set(client_id=client_id, limit=limit)
    if df.empty:
        return {"error": "no labeled data"}
    return {
        "n": len(df),
        "mae_kg": round(float(df["abs_error_kg"].mean()), 4),
        "slab_accuracy": round(float(df["slab_match"].mean()), 4),
        "basis_counts": df["basis"].value_counts().to_dict(),
    }


@router.get("/backtest/package")
def backtest_package(client_id: str, limit: int = 500):
    from engine.package import backtest
    df = backtest(client_id=client_id, limit=limit)
    if df.empty:
        return {"error": "no data"}
    return {
        "n": len(df),
        "original_slab_match_rate": round(float(df["slab_match_original"].mean()), 4),
        "recommended_slab_match_rate": round(float(df["slab_match_recommended"].mean()), 4),
        "decision_counts": df["decision"].value_counts().to_dict(),
    }
